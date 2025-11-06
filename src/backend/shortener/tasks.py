import os
import logging
import base64
import zipfile
import qrcode
import xlsxwriter
from io import BytesIO
from qrcode.image.svg import SvgPathImage
from qrcode.image.pil import PilImage
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from openpyxl import load_workbook
from celery import shared_task
from django.conf import settings
from django.core.files.storage import default_storage
from common.utils.chunk_parsing import get_chunks
from common.utils.extract_url_from_cell import extract_url_from_cell
from .models import Link



logger = logging.getLogger(__name__)

@shared_task(bind=True)
def generate_export_file(self, base_url, generate_qr=True):
    '''Создание и экспортирование ZIP файла, с Excel SVG|PDF файлами'''
    try:
        self.update_state(
            state='PROGRESS',
            meta={
                'current': 0,
                'total': 100, 
                'status': 'Подготовка данных'
            }
        )
        
        # Общее количествво ссылок, для progress bar-а
        total_links = Link.objects.count()
        processed_links = 0
        
        # Создание Excel файла
        excel_buffer = BytesIO()
        workbook = xlsxwriter.Workbook(excel_buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        headers = ['url', 'short_url', 'description', '']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        links = Link.objects.all()
        row = 1
        links_data = []
        
        # Обработка ссылок с обновлением прогресса
        for chunk in get_chunks(links, chunk_size=100):
            for link in chunk:
                links_data.append((link.url, link.code, link.description or ''))
                worksheet.write(row, 0, link.url)
                worksheet.write(row, 1, f"{base_url}{link.code}")
                worksheet.write(row, 2, link.description or '')
                row += 1
                processed_links += 1
                
                if processed_links % 10 == 0 and total_links > 0:
                    progress = int((processed_links / total_links) * 50)
                    self.update_state(
                        state='PROGRESS', 
                        meta={
                            'current': progress,
                            'total': 100,
                            'status': f'Обработано {processed_links} из {total_links} ссылок'
                        }
                    )

        workbook.close()
        excel_buffer.seek(0)
        
        self.update_state(
            state='PROGRESS',
            meta={
                'current': 50,
                'total': 100,
                'status': 'Создание ZIP архива'
            }
        )

        # Создание ZIP файла
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr('links.xlsx', excel_buffer.getvalue())

            if generate_qr:
                total_qr_codes = len(links_data)
                processed_qr = 0
                
                for url, code, description in links_data:
                    short_url = f"{base_url}{code}"

                    # Создание SVG QR-кодов
                    qr = qrcode.QRCode(version=1, box_size=10, border=4)
                    qr.add_data(short_url)
                    qr.make(fit=True)
                    svg_buffer = BytesIO()
                    qr.make_image(image_factory=SvgPathImage).save(svg_buffer)
                    zip_file.writestr(f"qr_codes/{code}.svg", svg_buffer.getvalue())

                    # Создание PNG QR-кодов
                    qr = qrcode.QRCode(version=1, box_size=10, border=4)
                    qr.add_data(short_url)
                    qr.make(fit=True)
                    png_buffer = BytesIO()
                    img = qr.make_image(image_factory=PilImage)
                    img.save(png_buffer, format="PNG")
                    zip_file.writestr(f"qr_codes/{code}.png", png_buffer.getvalue())

                    # Создание PDF QR-кодов
                    pdf_buffer = BytesIO()
                    img_width, img_height = img.size
                    
                    c = canvas.Canvas(pdf_buffer, pagesize=(img_width, img_height))
                    png_buffer.seek(0) 
                    image_reader = ImageReader(png_buffer)
                    
                    c.drawImage(image_reader, 0, 0, width=img_width, height=img_height)
                    c.showPage()
                    c.save()
                    
                    zip_file.writestr(f"qr_codes/{code}.pdf", pdf_buffer.getvalue())

                    processed_qr += 1
                    # Обновление прогресса для QR-кодов (оставшиеся 50%)
                    if total_qr_codes > 0:
                        qr_progress = 50 + int((processed_qr / total_qr_codes) * 50)
                        self.update_state(
                            state='PROGRESS', 
                            meta={
                                'current': qr_progress,
                                'total': 100,
                                'status': f'Создано {processed_qr} из {total_qr_codes} QR-кодов'
                            }
                        )

        zip_buffer.seek(0)
        
        # Сохранение файла на серваке
        filename = f"links_export_{self.request.id}.zip"
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        
        # Создание директории, в случае если она не существует
        os.makedirs(export_dir, exist_ok=True)
        
        filepath = os.path.join(export_dir, filename)
        
        with open(filepath, 'wb') as f:
            f.write(zip_buffer.getvalue())
        
        return {
            "file_path": filepath,
            "filename": filename,
            "download_url": f"/media/exports/{filename}",
            "file_size": os.path.getsize(filepath)
        }
        
    except Exception as e:
        self.update_state(state='FAILURE', meta={'error': str(e)})
        return {"error": str(e)}

@shared_task(bind=True)
def bulk_create_links(self, file_content_base64, base_url):
    '''Экспортирование Excel файла в БД'''
    try:
        # Декодирование содержимого файла из base64
        file_content = base64.b64decode(file_content_base64)
        file_buffer = BytesIO(file_content)
        
        # Чтение Excel файла
        workbook = load_workbook(file_buffer)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)) if cell.value]
        
        if 'url' not in headers:
            raise ValueError("В Excel файле должна быть колонка -> url")
        
        total_rows = sheet.max_row - 1
        
        self.update_state(
            state='PROGRESS',
            meta={
                'current': 0,
                'total': total_rows,
                'percent': 0,
                'stage': 'Подсчет строк'
            }
        )

        links_to_create = []
        created_links = []
        batch_size = 1000
        processed_rows = 0

        # Получаем индексы колонок
        url_col_index = headers.index('url')
        description_col_index = headers.index('description') if 'description' in headers else -1
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            url_cell = row[url_col_index]
            url = extract_url_from_cell(url_cell)
            
            # Пропускаем строки без валидного URL
            if not url or not url.startswith('http'):
                processed_rows += 1
                continue
            
            # Создаем данные для ссылки
            link_data = {"url": url}
            if description_col_index != -1:
                description_cell = row[description_col_index]
                link_data["description"] = description_cell.value or ''
            
            links_to_create.append(Link(**link_data))
            processed_rows += 1
            
            # Обновляем прогресс каждые 10 строк
            if processed_rows % 10 == 0:
                percent = (processed_rows / total_rows) * 100
                self.update_state(
                    state='PROGRESS',
                    meta={
                        'current': processed_rows,
                        'total': total_rows,
                        'percent': round(percent, 2),
                        'stage': 'Обработка строк'
                    }
                )

            if len(links_to_create) >= batch_size:
                self.update_state(
                    state='PROGRESS',
                    meta={
                        'current': processed_rows,
                        'total': total_rows,
                        'percent': round((processed_rows / total_rows) * 100, 2),
                        'stage': 'Сохраняем в БД'
                    }
                )
                Link.objects.bulk_create(links_to_create, batch_size=batch_size)
                for link in links_to_create:
                    code = f"{base_url}{link.code}"
                    created_links.append({
                        "url": link.url,
                        "code": code,
                        "description": link.description,
                    })
                links_to_create = []

        # Создание оставшихся записей
        if links_to_create:
            self.update_state(
                state='PROGRESS',
                meta={
                    'current': total_rows,
                    'total': total_rows,
                    'percent': 100,
                    'stage': 'Сохранение последних записей в БД'
                }
            )
            Link.objects.bulk_create(links_to_create, batch_size=batch_size)
            for link in links_to_create:
                code = f"{base_url}{link.code}"
                created_links.append({
                    "url": link.url,
                    "code": code,
                    "description": link.description,
                })

        workbook.close()
        
        return {
            "created_links": created_links,
            "total_processed": processed_rows,
            "total_created": len(created_links)
        }
    
    except Exception as e:
        self.update_state(
            state='FAILURE',
            meta={'error': str(e)}
        )
        raise